"""
Módulo de processamento de arquivos para a aplicação de Indexação de Documentos.
Contém funções para extrair texto de diferentes tipos de arquivos.
"""

import os
import logging
from typing import Dict, Any, Optional

import pandas as pd
import openpyxl
import xlrd
from docx import Document
from PyPDF2 import PdfReader

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def extract_text_from_docx(file_path: str) -> str:
    """
    Extrai texto de arquivos DOCX.

    Args:
        file_path: Caminho para o arquivo DOCX

    Returns:
        Texto extraído como uma string
    """
    try:
        doc = Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        logger.error(f"Erro ao extrair texto do arquivo DOCX {file_path}: {str(e)}")
        return ""

def extract_text_from_xlsx(file_path: str) -> str:
    """
    Extrai texto de arquivos XLSX.

    Args:
        file_path: Caminho para o arquivo XLSX

    Returns:
        Texto extraído como uma string
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value:
                        row_text.append(str(cell.value))
                if row_text:
                    text.append(" ".join(row_text))
        return "\n".join(text)
    except Exception as e:
        logger.error(f"Erro ao extrair texto do arquivo XLSX {file_path}: {str(e)}")
        return ""

def extract_text_from_xls(file_path: str) -> str:
    """
    Extrai texto de arquivos XLS.

    Args:
        file_path: Caminho para o arquivo XLS

    Returns:
        Texto extraído como uma string
    """
    try:
        wb = xlrd.open_workbook(file_path)
        text = []
        for sheet in wb.sheet_names():
            ws = wb.sheet_by_name(sheet)
            for row in range(ws.nrows):
                row_text = []
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(row, col)
                    if cell_value:
                        row_text.append(str(cell_value))
                if row_text:
                    text.append(" ".join(row_text))
        return "\n".join(text)
    except Exception as e:
        logger.error(f"Erro ao extrair texto do arquivo XLS {file_path}: {str(e)}")
        return ""

def extract_text_from_pdf(file_path: str) -> str:
    """
    Extrai texto de arquivos PDF.

    Args:
        file_path: Caminho para o arquivo PDF

    Returns:
        Texto extraído como uma string
    """
    try:
        reader = PdfReader(file_path)
        text = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)
        return "\n".join(text)
    except Exception as e:
        logger.error(f"Erro ao extrair texto do arquivo PDF {file_path}: {str(e)}")
        return ""

def extract_text_from_csv(file_path: str) -> str:
    """
    Extrai texto de arquivos CSV.

    Args:
        file_path: Caminho para o arquivo CSV

    Returns:
        Texto extraído como uma string
    """
    try:
        df = pd.read_csv(file_path)
        return df.to_string(index=False)
    except Exception as e:
        logger.error(f"Erro ao extrair texto do arquivo CSV {file_path}: {str(e)}")
        return ""

def process_file(file_path: str) -> Optional[Dict[str, Any]]:
    """
    Processa um arquivo e extrai seu conteúdo com base no tipo de arquivo.

    Args:
        file_path: Caminho para o arquivo

    Returns:
        Dicionário com metadados e conteúdo do arquivo, ou None se o processamento falhar
    """
    try:
        file_name = os.path.basename(file_path)
        file_ext = os.path.splitext(file_name)[1].lower()
        content = ""

        if file_ext == '.docx':
            content = extract_text_from_docx(file_path)
        elif file_ext == '.xlsx':
            content = extract_text_from_xlsx(file_path)
        elif file_ext == '.xls':
            content = extract_text_from_xls(file_path)
        elif file_ext == '.pdf':
            content = extract_text_from_pdf(file_path)
        elif file_ext == '.csv':
            content = extract_text_from_csv(file_path)
        elif file_ext == '.doc':
            # Para arquivos DOC, você pode precisar de ferramentas adicionais
            logger.warning(f"Formato de arquivo DOC tem suporte limitado: {file_path}")
            content = "Texto de arquivo DOC - considere usar antiword para extração completa"
        else:
            logger.warning(f"Tipo de arquivo não suportado: {file_ext}")
            return None

        if not content:
            logger.warning(f"Nenhum conteúdo extraído de {file_path}")
            return None

        return {
            "file_name": file_name,
            "file_path": file_path,
            "file_type": file_ext[1:],  # Remove o ponto
            "content": content,
            "file_size": os.path.getsize(file_path)
        }
    except Exception as e:
        logger.error(f"Erro ao processar arquivo {file_path}: {str(e)}")
        return None
